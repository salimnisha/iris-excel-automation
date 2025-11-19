from openpyxl.styles import PatternFill, Font, Border, Side

# Files
SOURCE_FILE = r"Ongoing Deliverable_US-1-US AU-1896858.1_Synopsys Inc._GDC EMSS PM Support_10.29.2025.xlsx"

# Rows and Columns where pivots start
PIVOT1_START_ROW = 1  # Overdue pivot
PIVOT2_START_ROW = 1  # Due date 1-14 days
PIVOT3_START_ROW = 1  # Count of content
PIVOT4_START_ROW = 1  # Status is Addressed
PIVOT5_START_ROW = 1  # Signoff Aging

PIVOT1_START_COL = 1  # Overdue pivot
PIVOT2_START_COL = 4  # Due date 1-14 days
PIVOT3_START_COL = 8  # Count of content
PIVOT4_START_COL = 11  # Status is Addressed
PIVOT5_START_COL = 14  # Signoff Aging


# Columns where summary tables start
TABLE1_START_COL = 1  # Open review notes table
TABLE2_START_COL = 9  # Addressed review notes table
TABLE3_START_COL = 14  # Signoff Aging table

# No. of lines to leave below the longest pivot table before starting the tables
BUFFER_LINES = 9

# Styles
HEADER_FILL = PatternFill(fill_type="solid", start_color="B8CCE4")
SECTION_FILL = PatternFill(fill_type="solid", start_color="DEE6F0")
THIN_BOTTOM = Border(bottom=Side(style="thin"))
THIN_TOP = Border(top=Side(style="thin"))
THIN_ALL_SIDES = Border(
    top=Side(style="thin"),
    bottom=Side(style="thin"),
    right=Side(style="thin"),
    left=Side(style="thin"),
)
RED = "FF0000"

# Source data for dataframe
DF1_SHEET = "ReviewNoteAging"
DF1_SHEET_HEADER = 6  # Data starts in excel row 7 => header 7

DF2_SHEET = "SignoffAging"
DF2_SHEET_HEADER = 6  # Data starts in excel row 7 => header 7


# Sheet names
CALC_SHEET = "Calculations"
REPORT_SHEET = "Report"
BASE_DATE_SHEET = "ReviewNoteAging"
BASE_DATE_CELL = "B4"
LAST_SYNC_SHEET = "SignoffAging"
LAST_SYNC_CELL = "B4"
