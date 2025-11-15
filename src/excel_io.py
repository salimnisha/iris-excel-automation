# ======================================
# IMPORTS
# ======================================
import shutil
import pandas as pd
import xlwings as xw
import re
from datetime import datetime
from openpyxl import load_workbook


def make_copy(source_file: str, debug: bool = False) -> str:
    """Copies source file to make a working copy with tabs for Calculations and Report"""

    copy_file = f"REPORT_{source_file}"
    shutil.copy(source_file, copy_file)

    # Create tabs
    wb = load_workbook(copy_file)
    tabs = ["Calculations", "REPORT"]
    for tab in tabs:
        if tab in wb.sheetnames:
            del wb[tab]
        wb.create_sheet(tab, index=tabs.index(tab))

    if debug:
        print("\n🐞 ====== DEBUG BLOCK START: make_copy (report_builder.py) ======")
        ws = wb["ReviewNoteAging"]
        print("[DEBUG] Checking VLOOKUP formula cells (Colum R-S, Row 7-12):")
        # Check if formulas survived
        # Expected: Some VLOOKUP formulas in list
        for row in ws.iter_rows(
            min_row=7, max_row=12, min_col=18, max_col=19, values_only=False
        ):
            print([cell.value for cell in row])
        print("🐞 ====== DEBUG BLOCK END: make_copy (report_builder.py) ====== \n")

    wb.save(copy_file)
    wb.close()  # Close the file so it does not mess with any other libraries using the same file later

    print("\n✅ Copied file to:", copy_file)
    return copy_file


def force_excel_recalc(filename: str):
    """Force complete recalculation of all formulas in the excel sheet

    Useful if any formulas were written programmatically, and the values need to be read by other libraries later
    """
    app = xw.App(visible=False)  # run in the background
    wb = app.books.open(filename)

    # Force recalculate all formulas
    wb.app.calculate()

    # Remember to save and quit, so an open file in the background does not mess with other libraries later
    wb.save()
    wb.close()
    app.quit()


def read_excel_dataframe(
    file_name: str, sheet_name: str, header_start: int, debug: bool = False
) -> pd.DataFrame:
    """Read the excel and return dataframe from the required sheet"""

    df = pd.read_excel(io=file_name, sheet_name=sheet_name, header=header_start)
    if debug:
        print("\n🐞 ====== DEBUG BLOCK START: get_dataframe (report_builder.py) ======")
        print(
            f"[DEBUG] Reading data from {file_name}"
        )  # Expected: Read from REPORT_{filename}
        print("🐞 ====== DEBUG BLOCK END: get_dataframe (report_builder.py) ====== \n")

    return df


def extract_base_date(ws, cell) -> datetime:
    """Extract date from the given sheet and cell"""
    cell_value = str(ws[cell].value)
    match = re.search(r"(\d{1,2}/\d{1,2}/\d{4})", cell_value)

    base_date = datetime.strptime(match.group(1), "%m/%d/%Y") if match else None

    if base_date is None:
        raise ValueError("⚠️ Base date not found!")

    print(
        f"📆 Base date extracted: {base_date.strftime('%Y-%m-%d') if base_date else '⚠️ Date Not Found'}"
    )

    return base_date
