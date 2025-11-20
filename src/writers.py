# ======================================
# IMPORTS
# ======================================
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

import src.constants as C
from src.formatting import autofit_colums
from src.excel_io import force_excel_recalc, load_values_only_workbook


def write_simple_pivot(ws, pivot_df, start_row, start_col, title=None, debug=False):
    """
    Write a pivot table to an Excel sheet
    at a specific location. Writes a simple flat table with 2 columns.

    Returns a dict of start and end row and column values
    {"start_row": 3, "end_row": 10, "start_col": 1, "end_col": 2}
    """

    # Initialize address to return
    pivot_address = {}

    # ----------------------------------------------------------------
    # 1. Write the header row(s)
    # ----------------------------------------------------------------
    # Title header
    if title:
        header_cell = ws.cell(row=start_row, column=start_col, value=title)
        header_cell.font = Font(bold=True)
        header_cell.fill = C.HEADER_FILL
        header_cell = ws.cell(row=start_row, column=start_col + 1, value="")
        header_cell.fill = C.HEADER_FILL

        start_row = start_row + 2  # Leave one row below before pivot data header

    # Write header
    value_col_name = pivot_df.columns.to_list()[0]
    header_cell = ws.cell(row=start_row, column=start_col, value="Row Labels")
    header_cell.font = Font(bold=True)
    header_cell.fill = C.SECTION_FILL
    header_cell = ws.cell(row=start_row, column=start_col + 1, value=value_col_name)
    header_cell.font = Font(bold=True)
    header_cell.fill = C.SECTION_FILL

    # Write data
    current_row = start_row + 1  # Row after header
    for index, row_val in pivot_df.iterrows():
        for j, col_name in enumerate(pivot_df.columns, start=start_col):
            ws.cell(row=current_row, column=j, value=index)
            ws.cell(row=current_row, column=j + 1, value=row_val[col_name])
            current_row += 1

    # Calculate grand_total and write to last row
    grand_total = pivot_df[value_col_name].sum()
    grand_cell = ws.cell(row=current_row, column=start_col, value="Total")
    grand_cell.font = Font(bold=True)
    grand_cell.fill = C.SECTION_FILL
    grand_cell.border = C.THIN_TOP

    grand_cell = ws.cell(row=current_row, column=start_col + 1, value=grand_total)
    grand_cell.font = Font(bold=True)
    grand_cell.fill = C.SECTION_FILL
    grand_cell.border = C.THIN_TOP

    # Set column width
    autofit_colums(ws, start_col=start_col, end_col=start_col + 1, limit_width=True)

    # Range occupied by the pivot table
    pivot_address["start_row"] = start_row
    pivot_address["start_col"] = start_col
    pivot_address["end_row"] = current_row
    pivot_address["end_col"] = start_col + 1  # this is our second header_cell

    return pivot_address


def write_multi_index_pivot(
    ws, pivot_df, start_row, start_col, title=None, debug=False
):
    """Writes a multi-index pivot to the worksheet
    - First index (main group) is written bold
    - Second index (items under main group) is indented under corresponding main group
    - Only one Values column

    Returns a dict of start and end row and column values
    {"start_row": 3, "end_row": 10, "start_col": 1, "end_col": 2}
    """

    # Check if dataframe is actually multi-index
    if not isinstance(pivot_df.index, pd.MultiIndex):
        raise ValueError("⚠️ Error: Pivot dataframe must have 2-level MultiIndex")

    level0_name, level1_name = pivot_df.index.names
    value_col_name = pivot_df.columns.to_list()[0]

    # Initialize address to return
    pivot_address = {}

    # ----------------------------------------------------------------
    # 1. Write the header row(s)
    # ----------------------------------------------------------------
    # Title header
    if title:
        header_cell = ws.cell(row=start_row, column=start_col, value=title)
        header_cell.font = Font(bold=True)
        header_cell.fill = C.HEADER_FILL
        header_cell = ws.cell(row=start_row, column=start_col + 1, value="")
        header_cell.fill = C.HEADER_FILL

        start_row = start_row + 2  # Leave one row below before pivot data header

    # Pivot data header
    header_cell = ws.cell(row=start_row, column=start_col, value="Row Labels")
    header_cell.font = Font(bold=True)
    header_cell.fill = C.SECTION_FILL
    header_cell = ws.cell(row=start_row, column=start_col + 1, value=value_col_name)
    header_cell.font = Font(bold=True)
    header_cell.fill = C.SECTION_FILL

    pivot_address["start_row"] = start_row
    pivot_address["start_col"] = start_col
    pivot_address["end_col"] = start_col + 1  # this is our second header_cell

    # ----------------------------------------------------------------
    # 2. Walk through the MultiIndex and write the rows
    # ----------------------------------------------------------------

    # Calculate the totals for the level0 groups ('Audit', 'TA').
    totals = pivot_df.groupby(level=0)[value_col_name].sum()
    grand_total = 0

    current_row = start_row + 1  # We start writing the row data from here
    current_group = None  # Track the main group, or level0 index ('Audit' or 'TA')

    for (assigned_group, allocated_to), row_value in pivot_df.iterrows():
        # 🤓 df.iterrows() lets you iterate over DataFrame rows as (index, Series) pairs.
        # index: label or tuple of labels; Series: data of the row as a series. Example below
        #   Index: ('TA', 'Anika Parkar')
        #   Row: Overdue    3

        # If assigned group changes, write bold group header
        if assigned_group != current_group:
            group_cell = ws.cell(
                row=current_row, column=start_col, value=assigned_group
            )
            group_cell.font = Font(bold=True)
            group_cell.border = C.THIN_BOTTOM

            # Write total for the group
            total = totals[assigned_group]
            total_cell = ws.cell(row=current_row, column=start_col + 1, value=total)
            total_cell.font = Font(bold=True)
            total_cell.border = C.THIN_BOTTOM

            grand_total += total

            current_group = assigned_group
            current_row += 1

        # Write second-level rows of allocated_to, indented below assigned_group
        allocated_to_cell = ws.cell(
            row=current_row, column=start_col, value=allocated_to
        )
        allocated_to_cell.alignment = Alignment(indent=1)
        # Write value for the allocated_to person, retrieved using name of the value column as key
        # e.g. row_value["Overdue"] = 3  -- row_value is a variable from the for loop
        ws.cell(row=current_row, column=start_col + 1, value=row_value[value_col_name])

        current_row += 1

    # Write the grand total at the very end
    grand_cell = ws.cell(row=current_row, column=start_col, value="Grand Total")
    grand_cell.font = Font(bold=True)
    grand_cell.fill = C.SECTION_FILL
    grand_cell.border = C.THIN_TOP

    grand_cell = ws.cell(row=current_row, column=start_col + 1, value=grand_total)
    grand_cell.font = Font(bold=True)
    grand_cell.fill = C.SECTION_FILL
    grand_cell.border = C.THIN_TOP

    # Set column width
    autofit_colums(ws, start_col=start_col, end_col=start_col + 1, limit_width=True)

    if debug:
        print(
            "\n🐞 ====== DEBUG BLOCK START: write_multi_index_pivot (writers.py) ======"
        )
        print("[DEBUG] Pivot title:", title)
        print("[DEBUG] Groups written:", pivot_df.index.levels[0].to_list())
        print("[DEBUG] Total rows:", current_row - start_row)
        print(
            "🐞 ====== DEBUG BLOCK END: write_multi_index_pivot (writers.py) ======\n"
        )

    pivot_address["end_row"] = current_row

    return pivot_address


def write_pivot_tables_to_sheet(pivots, ws, debug=False):
    # -----------------------------------------------------
    # Write the pivot tables to Calculations tab
    # -----------------------------------------------------
    ws_calc = ws
    overdue_pivot = pivots["overdue"]
    due_date_pivot = pivots["due_date"]
    count_of_content_pivot = pivots["count_of_content"]
    addressed_status_pivot = pivots["addressed_status"]
    signoff_aging_pivot = pivots["signoff_aging"]

    pivots_ranges = {
        "overdue": {},
        "due_date": {},
        "count_of_content": {},
        "addressed_status": {},
        "signoff_aging": {},
    }

    # Pivot1: Write Overdue pivot
    title = "Filter: 'Aged' > 0"
    address = write_multi_index_pivot(
        ws=ws_calc,
        pivot_df=overdue_pivot,
        start_row=C.PIVOT1_START_ROW,
        start_col=C.PIVOT1_START_COL,
        title=title,
        debug=debug,
    )
    pivots_ranges["overdue"] = address
    print(
        "\n✅ 1. Overdue pivot written to Calculations tab up to row",
        address["end_row"],
    )

    # Pivot2: Write Due within 1-14 days pivot
    title = "Filter Applied: 'Due Date' 1-14 days (inclusive of start date)"
    address = write_multi_index_pivot(
        ws=ws_calc,
        pivot_df=due_date_pivot,
        start_row=C.PIVOT2_START_ROW,
        start_col=C.PIVOT2_START_COL,
        title=title,
        debug=debug,
    )
    pivots_ranges["due_date"] = address
    print(
        "\n✅ 2. Due Date pivot written to Calculations tab up to row",
        address["end_row"],
    )

    # Pivot3: Write Count of Content
    title = "Filter: None Applied"
    address = write_multi_index_pivot(
        ws=ws_calc,
        pivot_df=count_of_content_pivot,
        start_row=C.PIVOT3_START_ROW,
        start_col=C.PIVOT3_START_COL,
        title=title,
        debug=debug,
    )
    pivots_ranges["count_of_content"] = address
    print(
        "\n✅ 3. Count of Content pivot written to Calculations tab up to row",
        address["end_row"],
    )

    # Pivot4: Write Status is Addressed
    title = "Filter: 'Status' == 'Addressed'"
    address = write_multi_index_pivot(
        ws=ws_calc,
        pivot_df=addressed_status_pivot,
        start_row=C.PIVOT4_START_ROW,
        start_col=C.PIVOT4_START_COL,
        title=title,
        debug=debug,
    )
    pivots_ranges["addressed_status"] = address
    print(
        "\n✅ 4. Addressed Status pivot written to Calculations tab up to row",
        address["end_row"],
    )

    # Pivot5: Signoff Aging [with Signoff Role != ('In-Charge' or 'Senior')]
    #   Note: This is a simple pivot with just one index, not a multi-index pivot
    #   So call the write_simple_pivot method
    title = "Signoff Role != ('In-Charge' or 'Senior')"
    address = write_simple_pivot(
        ws=ws_calc,
        pivot_df=signoff_aging_pivot,
        start_row=C.PIVOT4_START_ROW,
        start_col=C.PIVOT5_START_COL,
        title=title,
        debug=debug,
    )
    pivots_ranges["signoff_aging"] = address
    print(
        "\n✅ 5. Signoff aging pivot written to Calculations tab up to row",
        address["end_row"],
    )

    return pivots_ranges


def write_table(ws, start_row, start_col, title, header, rows, row_border=False):
    """Write a simple table into an Openpyxl worksheet

    Args:
        ws (Worksheet): Worksheet object to write
        start_row (int): Top left row to place title
        start_col (int): Top left column to place title
        title (str): Title above the header
        header (list[str]): list of column names
        rows (list[list)]: Table data as list of rows
    """

    # Table address to return
    table_address = {}

    # Write the title
    cell = ws.cell(row=start_row, column=start_col, value=title)
    cell.font = Font(bold=True, color=C.RED)

    table_address["start_row"] = start_row
    table_address["start_col"] = start_col

    # Write the column header one row below
    header_row = start_row + 1
    for j, header in enumerate(header):
        cell = ws.cell(row=header_row, column=start_col + j, value=header)
        cell.font = Font(bold=True)
        cell.border = C.THIN_ALL_SIDES

    last_row = 0
    last_col = 0
    # Write table contents
    data_start_row = start_row + 2
    for i_row, row in enumerate(rows):
        for j_col, value in enumerate(row):
            cell = ws.cell(
                row=data_start_row + i_row, column=start_col + j_col, value=value
            )
            if row_border:
                cell.border = C.THIN_ALL_SIDES
            last_row = data_start_row + i_row
            last_col = start_col + j_col

    table_address["end_row"] = last_row
    table_address["end_col"] = last_col

    return table_address


def write_summary_tables_to_sheet(tables, ws, start_row, debug=False):
    # -----------------------------------------------------
    # Write the summary tables to Calculations tab
    # -----------------------------------------------------
    ws_calc = ws
    open_notes_table = tables["open_notes"]
    addressed_notes_table = tables["addressed_notes"]
    signoff_aging_table = tables["signoff_aging"]

    table_ranges = {
        "open_notes": {},
        "addressed_notes": {},
    }

    # Table1: Write Open Notes Summary table
    title = open_notes_table["title"]
    header = open_notes_table["header"]
    rows = open_notes_table["rows"]
    address = write_table(
        ws=ws_calc,
        start_row=start_row,
        start_col=C.TABLE1_START_COL,  # Starts at first column
        title=title,
        header=header,
        rows=rows,
    )
    table_ranges["open_notes"] = address

    # Table2: Write Addressed Notes Summary table
    title = addressed_notes_table["title"]
    header = addressed_notes_table["header"]
    rows = addressed_notes_table["rows"]
    address = write_table(
        ws=ws_calc,
        start_row=start_row,
        start_col=C.TABLE2_START_COL,  # starts at 9th column
        title=title,
        header=header,
        rows=rows,
    )
    table_ranges["addressed_notes"] = address

    # Table3: Write Signoff Aging Table
    title = signoff_aging_table["title"]
    header = signoff_aging_table["header"]
    rows = signoff_aging_table["rows"]
    address = write_table(
        ws=ws_calc,
        start_row=start_row,
        start_col=C.TABLE3_START_COL,
        title=title,
        header=header,
        rows=rows,
    )
    table_ranges["signoff_aging"] = address

    return table_ranges


def copy_range_values_only(
    ws_src,
    ws_dst,
    src_start_row,
    src_start_col,
    height,
    width,
    dst_start_row,
    dst_start_col,
):
    """Copies a rectangular block of cells (values only) from one Excel sheet to another
    Does NOT copy formulas or formatting.

    Args:
        ws_src (Worksheet): Source worksheet
        ws_dst (Worksheet): Destination worksheet
        src_start_row (int): source range start row
        src_start_col (int): source range end row
        height (int): source range height (rows)
        width (int): source range width (columns)
        dst_start_row (int): destination range start row
        dst_start_col (int): destination range end row
    """

    # Destination range to return
    dst_range = {"start_row": dst_start_row, "start_col": dst_start_col}

    for r in range(height):
        for c in range(width):
            src_cell = ws_src.cell(row=src_start_row + r, column=src_start_col + c)
            dst_cell = ws_dst.cell(row=dst_start_row + r, column=dst_start_col + c)

            # Copy 'Values Only'
            dst_cell.value = src_cell.value

        dst_range["end_col"] = dst_start_col + c
    dst_range["end_row"] = dst_start_row + r

    return dst_range


def copy_table_to_report(
    src_path, wb_src, table_range, report_start_row, report_start_col, debug=False
):
    """Copy the calculated table range from source sheet (Calculations) to destination sheet (Report), given the table range

    Args:
        ws_calc (Worksheet): worksheet
        ws_report (Worksheet): worksheet
        table_range (dict): Dict of table range {start_row, start_col, end_row, end_col}
        dst_start_row (int): destination sheet start row
        dst_start_col (int): destination sheet start column
    """

    # Open a data_only mode spreadsheet to read the values (not formulas)
    wb_src.save(src_path)
    force_excel_recalc(src_path)

    wb_values = load_values_only_workbook(src_path)

    table_height = table_range["end_row"] - table_range["start_row"] + 1
    table_width = table_range["end_col"] - table_range["start_col"] + 1

    # The values are read from a new data_only worksheet, and written to the ws_report (passed via argument)
    # ⚠️ Don't save the data_only sheet (it will remove all formulas)
    report_range = copy_range_values_only(
        ws_src=wb_values[C.CALC_SHEET],
        ws_dst=wb_src[C.REPORT_SHEET],
        src_start_row=table_range["start_row"],
        src_start_col=table_range["start_col"],
        height=table_height,
        width=table_width,
        dst_start_row=report_start_row,
        dst_start_col=report_start_col,
    )

    return report_range


def copy_all_tables_to_report(file_path, wb_src, table_ranges, debug=False):
    """Copy all tables created in Calculations sheet to the Report sheet for formatting

        Returns a dict of report ranges

    Args:
        src_path (str): full path of sourcefile with extension
        wb_src (Openpyxl Workbook): Main workbook that the reports are written into
        table_ranges (dict): Table co-ordinates
        debug (bool, optional): Debug print or not. Defaults to False.
    """

    # Unpack table dict
    open_notes_table = table_ranges["open_notes"]  # Table 1
    addressed_notes_table = table_ranges["addressed_notes"]  # Table 2
    signoff_aging_table = table_ranges["signoff_aging"]  # Table 3

    # Initialize report range dict to return
    report_ranges = {}

    # Write each table
    # Table 1: Open Review Notes table
    open_notes_start_row = C.REPORT1_START_ROW
    open_notes_start_col = C.REPORT1_START_COL

    if open_notes_table:
        r1_range = copy_table_to_report(
            src_path=file_path,
            wb_src=wb_src,
            table_range=open_notes_table,
            report_start_row=open_notes_start_row,
            report_start_col=open_notes_start_col,
        )
        report_ranges["open_notes"] = r1_range

    # Table 2: Addressed Review Notes table
    addressed_notes_start_row = C.REPORT2_START_ROW
    addressed_notes_start_col = C.REPORT2_START_COL

    if addressed_notes_table:
        r2_range = copy_table_to_report(
            src_path=file_path,
            wb_src=wb_src,
            table_range=addressed_notes_table,
            report_start_row=addressed_notes_start_row,
            report_start_col=addressed_notes_start_col,
        )
        report_ranges["addressed_notes"] = r2_range

    # Table 1: Signoff Aging table
    signoff_aging_start_row = C.REPORT3_START_ROW
    signoff_aging_start_col = C.REPORT3_START_COL

    if signoff_aging_table:
        r3_range = copy_table_to_report(
            src_path=file_path,
            wb_src=wb_src,
            table_range=signoff_aging_table,
            report_start_row=signoff_aging_start_row,
            report_start_col=signoff_aging_start_col,
        )
        report_ranges["signoff_aging_notes"] = r3_range

    return report_ranges
