from openpyxl.styles import PatternFill, Border, Font, Side, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

import src.constants as C

# ========================================================
# BASIC COLORS AND FONTS
# ========================================================
RED = "FF0000"
WHITE = "FFFFFF"
DARK_BLUE = "4F81BD"
BLACK = "000000"
GREY = "D9D9D9"
SLATE_GREY = "768692"
DARK_GREY = "222222"

BOLD_FONT = Font(bold=True)

# Borders
INNER_CELL_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="dotted"),
    top=Side(style="dotted"),
    bottom=Side(style="dotted"),
)
LEFT_EDGE_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="dotted"),
    top=Side(style="dotted"),
    bottom=Side(style="dotted"),
)
TOP_EDGE_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="dotted"),
    top=Side(style="thin"),
    bottom=Side(style="dotted"),
)
RIGHT_EDGE_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="thin"),
    top=Side(style="dotted"),
    bottom=Side(style="dotted"),
)
BOTTOM_EDGE_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="dotted"),
    top=Side(style="dotted"),
    bottom=Side(style="thin"),
)
# Corners
TOP_LEFT_CORNER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="dotted"),
    top=Side(style="thin"),
    bottom=Side(style="dotted"),
)
TOP_RIGHT_CORNER_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="dotted"),
)
BOTTOM_LEFT_CORNER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="dotted"),
    top=Side(style="dotted"),
    bottom=Side(style="thin"),
)
BOTTOM_RIGHT_CORNER_BORDER = Border(
    left=Side(style="dotted"),
    right=Side(style="thin"),
    top=Side(style="dotted"),
    bottom=Side(style="thin"),
)

# Conditional formatting colors and fills
LIGHT_RED = "ff6c6c"
ORANGE = "ff9966"
DARK_YELLOW = "febf00"
LIGHT_YELLOW = "fdf099"
LIGHT_GREEN = "91d04f"
DARK_GREEN = "05b050"

PINK = "ffc7cd"
PALE_GREEN = "c6efcd"

# [ ] TODO: For conditional formatting, Patternfill doesn't seem to work without end_color. Check why
LIGHT_RED_FILL = PatternFill(
    fill_type="solid", start_color=LIGHT_RED, end_color=LIGHT_RED
)
ORANGE_FILL = PatternFill(fill_type="solid", start_color=ORANGE, end_color=ORANGE)
DARK_YELLOW_FILL = PatternFill(
    fill_type="solid", start_color=DARK_YELLOW, end_color=DARK_YELLOW
)
LIGHT_YELLOW_FILL = PatternFill(
    fill_type="solid", start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW
)
LIGHT_GREEN_FILL = PatternFill(
    fill_type="solid", start_color=LIGHT_GREEN, end_color=LIGHT_GREEN
)
DARK_GREEN_FILL = PatternFill(
    fill_type="solid", start_color=DARK_GREEN, end_color=DARK_GREEN
)
PINK_FILL = PatternFill(fill_type="solid", start_color=PINK, end_color=PINK)
PALE_GREEN_FILL = PatternFill(
    fill_type="solid", start_color=PALE_GREEN, end_color=PALE_GREEN
)

# Autofit Column dimensions
MAX_COL_WIDTH = 30
COLUMN_PADDING = 2

# Title, Header, and Footer styles
TITLE_FONT = Font(bold=True, color=RED)

HEADER_FONT = Font(bold=True, color=WHITE)
HEADER_FILL = PatternFill(fill_type="solid", start_color=DARK_BLUE)
HEADER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FOOTER_FONT = Font(bold=True, color=WHITE)
FOOTER_FILL = PatternFill(fill_type="solid", start_color=DARK_GREY)
FOOTER_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="double"),
)

GROUP_ROW_FILL = PatternFill(fill_type="solid", start_color=GREY)


# ========================================================
# BASIC FORMATTING
# ========================================================


def autofit_colums(ws, start_col, end_col, padding=COLUMN_PADDING, limit_width=True):
    """Set column width based on maximum length of content inside column"""
    for col in range(start_col, end_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
                if limit_width:
                    # max width should be <= MAX_COL_WIDTH
                    max_len = min(MAX_COL_WIDTH, max_len)
        ws.column_dimensions[col_letter].width = max_len + padding


def format_title(ws, row, col):
    """Style a header cell with bold font and title color"""
    cell = ws.cell(row=row, column=col)
    cell.font = TITLE_FONT


def format_header(ws, start_row, start_col, end_col):
    """Style a header with bold font, fill color"""
    num_cols = end_col - start_col + 1
    for col in range(num_cols):
        cell = ws.cell(row=start_row, column=start_col + col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER


def format_footer(ws, end_row, start_col, end_col):
    """Style the last row with bold font and fill color"""
    num_cols = end_col - start_col + 1
    for col in range(num_cols):
        cell = ws.cell(row=end_row, column=start_col + col)
        cell.font = FOOTER_FONT
        cell.fill = FOOTER_FILL
        cell.border = FOOTER_BORDER


def format_group_rows(
    ws, start_row, start_col, end_row, end_col, group_col, group_list
):
    """Fill the entire Group row with a grey colour.
    - group_col is the column number where the group names appear
    - group_list is a list of group names to check
    """
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=group_col)
        if cell.value in group_list:
            # Colour the entire row
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = GROUP_ROW_FILL
                cell.font = BOLD_FONT


def indent_child_rows(ws, start_row, end_row, group_col, group_list):
    """Indent all items appearing below the groups (from group_list)
    - group_col is the column with the groups and children
    """
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=group_col)
        if cell.value not in group_list and cell.value not in (None, ""):
            cell.alignment = Alignment(indent=2)


def format_table_data_cells(ws, start_row, start_col, end_row, end_col, debug=False):
    """Style the cells inside the table
    Values passed exclude title, header, footer"""

    # --------------------------------------------
    # 1. Set all the corners
    cell = ws.cell(row=start_row, column=start_col)
    cell.border = TOP_LEFT_CORNER_BORDER
    if debug:
        print(f"TOP_LEFT_CORNER_BORDER: ({start_row}, {start_col}), {cell.value}")

    cell = ws.cell(row=start_row, column=end_col)
    cell.border = TOP_RIGHT_CORNER_BORDER
    if debug:
        print(f"TOP_RIGHT_CORNER_BORDER: ({start_row}, {end_col})), {cell.value}")

    cell = ws.cell(row=end_row, column=start_col)
    cell.border = BOTTOM_LEFT_CORNER_BORDER
    if debug:
        print(f"BOTTOM_LEFT_CORNER_BORDER: ({end_row}, {start_col}), {cell.value}")

    cell = ws.cell(row=end_row, column=end_col)
    cell.border = BOTTOM_RIGHT_CORNER_BORDER
    if debug:
        print(f"BOTTOM_RIGHT_CORNER_BORDER: ({end_row}, {end_col}), {cell.value}")

    # --------------------------------------------
    # 2. Set all edges (excluding corners)
    # Top edge
    for col in range(start_col + 1, end_col):
        cell = ws.cell(start_row, col)
        cell.border = TOP_EDGE_BORDER
        if debug:
            print(f"TOP_EDGE_BORDER: ({start_row}, {col}), {cell.value}")

    # Bottom edge
    for col in range(start_col + 1, end_col):
        cell = ws.cell(end_row, col)
        cell.border = BOTTOM_EDGE_BORDER
        if debug:
            print(f"BOTTOM_EDGE_BORDER: ({end_row}, {col}), {cell.value}")

    # Left edge
    for row in range(start_row + 1, end_row):
        cell = ws.cell(row, start_col)
        cell.border = LEFT_EDGE_BORDER
        if debug:
            print(f"LEFT_EDGE_BORDER: ({row}, {start_col}), {cell.value}")

    # Right edge
    for row in range(start_row + 1, end_row):
        cell = ws.cell(row, end_col)
        cell.border = RIGHT_EDGE_BORDER
        if debug:
            print(f"RIGHT_EDGE_BORDER: ({row}, {end_col}), {cell.value}")

    # --------------------------------------------
    # 3. Inner cells
    for row in range(start_row + 1, end_row):
        for col in range(start_col + 1, end_col):
            cell = ws.cell(row, col)
            cell.border = INNER_CELL_BORDER
            if debug:
                print(f"INNER_CELL_BORDER: ({row}, {col}), {cell.value}")


def conditional_format_number_5color_scale(ws, cell_range):
    """Conditional format a number range"""

    # First item with stopIfTrue=True avoids formatting empty cells
    rules = [
        CellIsRule(operator="equal", formula=['""'], stopIfTrue=True),
        CellIsRule(operator="greaterThanOrEqual", formula=["30"], fill=LIGHT_RED_FILL),
        CellIsRule(operator="between", formula=["20", "29"], fill=ORANGE_FILL),
        CellIsRule(operator="between", formula=["10", "19"], fill=DARK_YELLOW_FILL),
        CellIsRule(operator="between", formula=["5", "9"], fill=LIGHT_YELLOW_FILL),
        CellIsRule(operator="between", formula=["2", "4"], fill=LIGHT_GREEN_FILL),
        CellIsRule(operator="between", formula=["0", "1"], fill=DARK_GREEN_FILL),
    ]

    for rule in rules:
        ws.conditional_formatting.add(cell_range, rule)


def conditional_format_number_positive_negative(ws, cell_range):
    """Conditional format for >0 and <0"""

    # First item with stopIfTrue=True avoids formatting empty cells
    rules = [
        CellIsRule(operator="equal", formula=['""'], stopIfTrue=True),
        CellIsRule(operator="greaterThan", formula=["0"], fill=PINK_FILL),
        CellIsRule(operator="lessThan", formula=["0"], fill=PALE_GREEN_FILL),
    ]

    for rule in rules:
        ws.conditional_formatting.add(cell_range, rule)


# ========================================================
# FORMATTING TABLES AS NEEDED
# ========================================================

# Generic formatting workflow
# 1. Apply conditional formatting
# 2. Format groups (if groups)
# 3. Format inside cells
# 4. Format header, footer, title


def apply_basic_formatting(ws_report, report_range):
    rep_start_row = report_range["start_row"]
    rep_start_col = report_range["start_col"]
    rep_end_row = report_range["end_row"]
    rep_end_col = report_range["end_col"]

    # Fit text to columns
    autofit_colums(ws=ws_report, start_col=rep_start_col, end_col=rep_end_col)

    # Format the title
    format_title(ws=ws_report, row=rep_start_row, col=rep_start_col)

    # Format the header rows
    header_row = rep_start_row + 1
    format_header(
        ws=ws_report,
        start_row=header_row,
        start_col=rep_start_col,
        end_col=rep_end_col,
    )

    # Set borders to all table data cells, and outside table border
    data_start_row = rep_start_row + 2  # First 2 rows are title and header
    data_end_row = rep_end_row - 1  # Last row is footer
    format_table_data_cells(
        ws=ws_report,
        start_row=data_start_row,
        start_col=rep_start_col,
        end_row=data_end_row,
        end_col=rep_end_col,
    )

    # Format the footer rows
    format_footer(
        ws=ws_report, end_row=rep_end_row, start_col=rep_start_col, end_col=rep_end_col
    )


def apply_indents_for_child_rows(ws_report, report_range, group_list):
    rep_start_row = report_range["start_row"]
    rep_start_col = report_range["start_col"]
    rep_end_row = report_range["end_row"]
    rep_end_col = report_range["end_col"]

    # Format the main group rows (Audit, TA) - color the entire row
    format_group_rows(
        ws=ws_report,
        start_row=rep_start_row,
        start_col=rep_start_col,
        end_row=rep_end_row,
        end_col=rep_end_col,
        group_col=rep_start_col,
        group_list=group_list,
    )

    # Indent the child rows - Indent all rows other than the main groups
    indent_start_row = rep_start_row + 2  # First 2 rows are title and header
    indent_end_row = rep_end_row - 1  # Last row is footer
    indent_child_rows(
        ws=ws_report,
        start_row=indent_start_row,
        end_row=indent_end_row,
        group_col=rep_start_col,
        group_list=group_list,
    )


def format_open_notes_report(ws_report, open_notes_range):
    """Format Report1: Open Review Notes (based on Open Review Notes table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """

    # ==== 1. Apply conditional formatting ====
    f_start_row = open_notes_range["start_row"] + 2  # First 2 rows are title and header
    f_end_row = open_notes_range["end_row"] - 1  # Last row is footer

    #   Conditional formatting #1 (5 scale) - Grand Total
    total_col = get_column_letter(C.REPORT1_GRAND_TOTAL_COL)
    total_cell_range = f"{total_col}{f_start_row}:{total_col}{f_end_row}"
    conditional_format_number_5color_scale(ws=ws_report, cell_range=total_cell_range)

    #   Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT1_DIFFERENCE_COL)
    diff_cell_range = f"{diff_col}{f_start_row}:{diff_col}{f_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # ==== 2. Indent child rows and fill-colour the group header rows ====
    assigned_groups = ["Audit", "TA"]
    apply_indents_for_child_rows(
        ws_report=ws_report, report_range=open_notes_range, group_list=assigned_groups
    )

    # ==== 3. Format table data cells (inner cells), header, footer, and title ====
    apply_basic_formatting(ws_report=ws_report, report_range=open_notes_range)

    print("\n⚑ 2. Formatted Open Review Notes Report")


# ---------------------------------------------------------


def format_addressed_notes_report(ws_report, addressed_notes_range):
    """Format Report1: Open Review Notes (based on Open Review Notes table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """

    # ==== 1. Apply conditional formatting ====
    f_start_row = (
        addressed_notes_range["start_row"] + 2
    )  # First 2 rows are title and header
    f_end_row = addressed_notes_range["end_row"] - 1  # Last row is footer

    # Conditional formatting #1 (5 scale) - Addressed column
    addr_col = get_column_letter(C.REPORT2_ADDRESSED_COL)
    addr_cell_range = f"{addr_col}{f_start_row}:{addr_col}{f_end_row}"
    conditional_format_number_5color_scale(ws=ws_report, cell_range=addr_cell_range)

    #   Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT2_DIFFERENCE_COL)
    diff_cell_range = f"{diff_col}{f_start_row}:{diff_col}{f_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # ==== 2. Indent child rows and fill-colour the group header rows ====
    assigned_groups = ["Audit", "TA"]
    apply_indents_for_child_rows(
        ws_report=ws_report,
        report_range=addressed_notes_range,
        group_list=assigned_groups,
    )

    # ==== 3. Format table data cells (inner cells), header, footer, and title ====
    apply_basic_formatting(ws_report=ws_report, report_range=addressed_notes_range)

    print("\n⚑ 1. Formatted Addressed Review Notes Report")


def format_signoff_aging_report(ws_report, signoff_aging_range):
    """Format Report3: Signoff Aging Report (based on Signoff Aging table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """
    # ==== 1. Apply conditional formatting ====
    f_start_row = (
        signoff_aging_range["start_row"] + 2
    )  # First 2 rows are title and header
    f_end_row = signoff_aging_range["end_row"] - 1  # Last row is footer

    #   Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT3_DIFFERENCE_COL)
    diff_cell_range = f"{diff_col}{f_start_row}:{diff_col}{f_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # ==== 2. Format table data cells (inner cells), header, footer, and title ====
    apply_basic_formatting(ws_report=ws_report, report_range=signoff_aging_range)

    print("\n⚑ 3. Signoff Aging Report")


def format_all_reports(ws_report, report_ranges):
    if report_ranges["open_notes"]:
        format_open_notes_report(
            ws_report=ws_report, open_notes_range=report_ranges["open_notes"]
        )

    if report_ranges["addressed_notes"]:
        format_addressed_notes_report(
            ws_report=ws_report, addressed_notes_range=report_ranges["addressed_notes"]
        )

    if report_ranges["signoff_aging_notes"]:
        format_signoff_aging_report(
            ws_report=ws_report,
            signoff_aging_range=report_ranges["signoff_aging_notes"],
        )


# -------------------------
# OLD - to be deleted
# -------------------------


def old_format_open_notes_report(ws_report, report_range):
    """Format Report1: Open Review Notes (based on Open Review Notes table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """

    rep_start_row = report_range["start_row"]
    rep_start_col = report_range["start_col"]
    rep_end_row = report_range["end_row"]
    rep_end_col = report_range["end_col"]

    # Fit text to columns
    autofit_colums(ws=ws_report, start_col=rep_start_col, end_col=rep_end_col)

    # Format the title
    format_title(ws=ws_report, row=rep_start_row, col=rep_start_col)

    # Format the header rows
    header_row = rep_start_row + 1
    format_header(
        ws=ws_report,
        start_row=header_row,
        start_col=rep_start_col,
        end_col=rep_end_col,
    )

    # Conditional formatting #1 (5 scale) - Grand Total
    total_col = get_column_letter(C.REPORT1_GRAND_TOTAL_COL)
    total_start_row = rep_start_row + 2  # First 2 rows are title and header
    total_end_row = rep_end_row - 1  # Last row is footer
    total_cell_range = f"{total_col}{total_start_row}:{total_col}{total_end_row}"
    conditional_format_number_5color_scale(ws=ws_report, cell_range=total_cell_range)

    # Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT1_DIFFERENCE_COL)
    diff_start_row = rep_start_row + 2  # First 2 rows are title and header
    diff_end_row = rep_end_row - 1  # Last row is footer
    diff_cell_range = f"{diff_col}{diff_start_row}:{diff_col}{diff_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # Format the main group rows (Audit, TA) - color the entire row
    main_groups = ["Audit", "TA"]
    format_group_rows(
        ws=ws_report,
        start_row=rep_start_row,
        start_col=rep_start_col,
        end_row=rep_end_row,
        end_col=rep_end_col,
        group_col=rep_start_col,
        group_list=main_groups,
    )

    # Indent the child rows - Indent all rows other than the main groups
    main_groups = ["Audit", "TA"]
    indent_start_row = rep_start_row + 2  # First 2 rows are title and header
    indent_end_row = rep_end_row - 1  # Last row is footer
    indent_child_rows(
        ws=ws_report,
        start_row=indent_start_row,
        end_row=indent_end_row,
        group_col=rep_start_col,
        group_list=main_groups,
    )

    # Set borders to all table data cells, and outside table border
    data_start_row = rep_start_row + 2  # First 2 rows are title and header
    data_end_row = rep_end_row - 1  # Last row is footer
    format_table_data_cells(
        ws=ws_report,
        start_row=data_start_row,
        start_col=rep_start_col,
        end_row=data_end_row,
        end_col=rep_end_col,
    )

    # Format the footer rows
    format_footer(
        ws=ws_report, end_row=rep_end_row, start_col=rep_start_col, end_col=rep_end_col
    )

    print("\n📍 1. Formatted Open Notes Report")


def old_format_addressed_notes_report(ws_report, report_range):
    """Format Report2: Addressed Review Notes (based on Addressed Review Notes table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """

    rep_start_row = report_range["start_row"]
    rep_start_col = report_range["start_col"]
    rep_end_row = report_range["end_row"]
    rep_end_col = report_range["end_col"]

    # Fit text to columns
    autofit_colums(ws=ws_report, start_col=rep_start_col, end_col=rep_end_col)

    # Format the title
    format_title(ws=ws_report, row=rep_start_row, col=rep_start_col)

    # Format the header rows
    header_row = rep_start_row + 1
    format_header(
        ws=ws_report,
        start_row=header_row,
        start_col=rep_start_col,
        end_col=rep_end_col,
    )
    print(
        f"Addressed notes Header -------- header_row: {header_row}, start_col: {rep_start_col}, end_col: {rep_end_col}"
    )

    # Conditional formatting #1 (5 scale) - Addressed column
    total_col = get_column_letter(C.REPORT2_ADDRESSED_COL)
    total_start_row = rep_start_row + 2  # First 2 rows are title and header
    total_end_row = rep_end_row - 1  # Last row is footer
    total_cell_range = f"{total_col}{total_start_row}:{total_col}{total_end_row}"
    conditional_format_number_5color_scale(ws=ws_report, cell_range=total_cell_range)

    # Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT2_DIFFERENCE_COL)
    diff_start_row = rep_start_row + 2  # First 2 rows are title and header
    diff_end_row = rep_end_row - 1  # Last row is footer
    diff_cell_range = f"{diff_col}{diff_start_row}:{diff_col}{diff_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # Format the main group rows (Audit, TA) - color the entire row
    main_groups = ["Audit", "TA"]
    format_group_rows(
        ws=ws_report,
        start_row=rep_start_row,
        start_col=rep_start_col,
        end_row=rep_end_row,
        end_col=rep_end_col,
        group_col=rep_start_col,
        group_list=main_groups,
    )

    # Indent the child rows - Indent all rows other than the main groups
    main_groups = ["Audit", "TA"]
    indent_start_row = rep_start_row + 2  # First 2 rows are title and header
    indent_end_row = rep_end_row - 1  # Last row is footer
    indent_child_rows(
        ws=ws_report,
        start_row=indent_start_row,
        end_row=indent_end_row,
        group_col=rep_start_col,
        group_list=main_groups,
    )

    # Set borders to all table data cells, and outside table border
    data_start_row = rep_start_row + 2  # First 2 rows are title and header
    data_end_row = rep_end_row - 1  # Last row is footer
    format_table_data_cells(
        ws=ws_report,
        start_row=data_start_row,
        start_col=rep_start_col,
        end_row=data_end_row,
        end_col=rep_end_col,
    )

    # Format the footer rows
    format_footer(
        ws=ws_report, end_row=rep_end_row, start_col=rep_start_col, end_col=rep_end_col
    )

    print("\n📍 2. Formatted Addressed Review Notes Report")


def old_format_signoff_aging_report(ws_report, report_range):
    """Format Report3: Signoff Aging Report (based on Signoff Aging table)

    Args:
        ws_report (Openpyxl Worksheet): handle to the report worksheet
        report_range (dict): start and end positions of the report
                            e.g. {"start_row": 1, "start_col": 1, "end_col": 7, "end_row": 38}
    """

    rep_start_row = report_range["start_row"]
    rep_start_col = report_range["start_col"]
    rep_end_row = report_range["end_row"]
    rep_end_col = report_range["end_col"]

    # Fit text to columns
    autofit_colums(ws=ws_report, start_col=rep_start_col, end_col=rep_end_col)

    # Format the title
    format_title(ws=ws_report, row=rep_start_row, col=rep_start_col)

    # Format the header rows
    header_row = rep_start_row + 1
    format_header(
        ws=ws_report,
        start_row=header_row,
        start_col=rep_start_col,
        end_col=rep_end_col,
    )

    # Conditional formatting #1 (> or < 0) - Difference
    diff_col = get_column_letter(C.REPORT3_DIFFERENCE_COL)
    diff_start_row = rep_start_row + 2  # First 2 rows are title and header
    diff_end_row = rep_end_row - 1  # Last row is footer
    diff_cell_range = f"{diff_col}{diff_start_row}:{diff_col}{diff_end_row}"
    conditional_format_number_positive_negative(
        ws=ws_report, cell_range=diff_cell_range
    )

    # Set borders to all table data cells, and outside table border
    data_start_row = rep_start_row + 2  # First 2 rows are title and header
    data_end_row = rep_end_row - 1  # Last row is footer
    format_table_data_cells(
        ws=ws_report,
        start_row=data_start_row,
        start_col=rep_start_col,
        end_row=data_end_row,
        end_col=rep_end_col,
    )

    # Format the footer rows
    format_footer(
        ws=ws_report, end_row=rep_end_row, start_col=rep_start_col, end_col=rep_end_col
    )

    print("\n📍 3. Formatted Signoff Aging Report")
